"""
Scrape building permits from Primaria Sector 1 website.
Downloads XLS files and parses them into permit data.
"""

import requests
import os
import tempfile
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime
import time

PAGE_URL = "https://primariasector1.ro/informatii-serviciul-urbanism/autorizatii-de-contruire-desfiintare/lista-autorizatiilor-de-construire/"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}


def _fetch_page_links():
    """Fetch all XLS/XLSX file links from the webpage"""
    try:
        response = requests.get(PAGE_URL, headers=HEADERS, timeout=30)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        links = []
        seen_urls = set()

        for a_tag in soup.find_all('a', href=True):
            href = a_tag['href']
            if href.endswith('.xls') or href.endswith('.xlsx'):
                if href in seen_urls:
                    continue
                seen_urls.add(href)

                text = a_tag.get_text(strip=True)
                original_filename = os.path.basename(href.split('?')[0])

                links.append({
                    'url': href,
                    'label': text,
                    'original_filename': original_filename,
                })

        return links

    except Exception as e:
        raise Exception(f"Error fetching page: {e}")


def _download_file(url, download_dir):
    """Download a file to the downloads directory"""
    try:
        response = requests.get(url, headers=HEADERS, timeout=60)
        response.raise_for_status()

        filename = os.path.basename(url.split('?')[0])
        filepath = download_dir / filename

        with open(filepath, 'wb') as f:
            f.write(response.content)

        return filepath

    except Exception:
        return None


def _detect_header_row(sheet, max_check=15):
    """Detect which row contains the headers"""
    header_keywords = [
        'nr', 'numar', 'număr', 'data', 'dată', 'adresa', 'adresă',
        'beneficiar', 'strada', 'stradă', 'ac', 'ad', 'autorizat',
        'lucrare', 'titular', 'solicitant', 'descriere', 'crt',
        'emiterii', 'inreg', 'cadastral', 'scopul'
    ]

    max_row_to_check = min(max_check, sheet.max_row)

    for row_idx in range(1, max_row_to_check + 1):
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

        matches = 0
        non_empty = 0

        for cell in row:
            if cell and str(cell).strip():
                non_empty += 1
                cell_lower = str(cell).lower().strip()
                if any(keyword in cell_lower for keyword in header_keywords):
                    matches += 1

        if matches >= 3 and non_empty > 0 and (matches / non_empty) >= 0.4:
            headers = []
            for cell in row:
                if cell:
                    header = str(cell).strip()
                    header = ' '.join(header.split())
                    headers.append(header)
                else:
                    headers.append(None)

            return row_idx, headers

    return None, None


def _find_address_column(headers):
    """Find which column contains the address"""
    for idx, header in enumerate(headers):
        if header and 'adresa' in header.lower():
            return idx
    return None


def _parse_file(filepath, file_url):
    """Parse a single XLS file and extract permits"""
    try:
        workbook = load_workbook(filepath, data_only=False)
        sheet = workbook.active

        header_row, headers = _detect_header_row(sheet)

        if header_row is None:
            return []

        address_col = _find_address_column(headers)
        if address_col is None:
            return []

        permits = []
        data_start_row = header_row + 1

        for row_idx in range(data_start_row, sheet.max_row + 1):
            row_cells = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]

            if not any(cell.value for cell in row_cells if cell.value):
                continue

            address_cell = row_cells[address_col] if address_col < len(row_cells) else None
            if not address_cell or not address_cell.value:
                continue

            address = str(address_cell.value).strip()

            data = {}
            for col_idx, (header, cell) in enumerate(zip(headers, row_cells)):
                if not header or not cell.value:
                    continue

                value = cell.value

                if cell.hyperlink:
                    value = cell.hyperlink.target
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')

                value_str = str(value).strip()
                data[header] = value_str

            permit = {
                'address': address,
                'data': data,
                'source': {
                    'issuer': 'ps1',
                    'url': file_url
                }
            }

            permits.append(permit)

        return permits

    except Exception:
        return []


def scrape_permits():
    """
    Scrape building permits from Primaria Sector 1.
    Returns list of permit dictionaries.
    """
    print("[PS1] Fetching file links from primariasector1.ro...")
    file_links = _fetch_page_links()

    if not file_links:
        raise Exception("No XLS files found on PS1 page")

    print(f"[PS1] Found {len(file_links)} XLS files to process")
    all_permits = []

    # Use temp directory for downloads
    with tempfile.TemporaryDirectory() as temp_dir:
        download_dir = Path(temp_dir)

        for idx, file_info in enumerate(file_links, 1):
            print(f"[PS1] [{idx}/{len(file_links)}] Processing: {file_info['original_filename']}")
            filepath = _download_file(file_info['url'], download_dir)

            if filepath:
                permits = _parse_file(filepath, file_info['url'])
                all_permits.extend(permits)
                print(f"[PS1]   -> Extracted {len(permits)} permits (total: {len(all_permits)})")

            time.sleep(0.3)

    print(f"[PS1] Done! Total permits extracted: {len(all_permits)}")
    return all_permits
